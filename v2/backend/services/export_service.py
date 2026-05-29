# -- Export service — generates Excel, CSV, and JSON from lead lists. --
# All files are written locally to the exports/ directory.

import csv
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from services.scraper_service import Lead

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path("../exports")
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = [
    "Name", "Category", "Address", "Phone",
    "Website", "Rating", "Review Count",
    "Search Query", "Location", "Scraped At",
]

ExportFormat = Literal["xlsx", "csv", "json"]

def _timestamped_name(prefix: str, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"

def export_leads(
    leads: list[Lead],
    fmt: ExportFormat = "xlsx",
    prefix: str = "leads",
) -> Path:
    filename = _timestamped_name(prefix, fmt)
    filepath = EXPORTS_DIR / filename

    if fmt == "xlsx":
        _write_xlsx(leads, filepath)
    elif fmt == "csv":
        _write_csv(leads, filepath)
    elif fmt == "json":
        _write_json(leads, filepath)
    else:
        raise ValueError(f"Unknown format: {fmt}")

    logger.info("Exported %d leads → %s", len(leads), filepath)
    return filepath

def _lead_row(lead: Lead) -> list:
    return [
        lead.name, lead.category, lead.address, lead.phone,
        lead.website, lead.rating, lead.review_count,
        lead.search_query, lead.location_query, lead.scraped_at,
    ]

def _write_xlsx(leads: list[Lead], path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header styling
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="F8FAFC", bold=True, size=11, name="Calibri")
    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    for row_num, lead in enumerate(leads, 2):
        for col_num, value in enumerate(_lead_row(lead), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_num % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    widths = [35, 25, 50, 18, 42, 8, 14, 35, 25, 22]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(path)

def _write_csv(leads: list[Lead], path: Path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for lead in leads:
            writer.writerow(_lead_row(lead))

def _write_json(leads: list[Lead], path: Path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump([lead.to_dict() for lead in leads], f, ensure_ascii=False, indent=2)

def list_exports() -> list[dict]:
    files = []
    for p in sorted(EXPORTS_DIR.glob("*"), key=os.path.getmtime, reverse=True):
        if p.suffix in {".xlsx", ".csv", ".json"}:
            files.append({
                "name": p.name,
                "size_kb": round(p.stat().st_size / 1024, 1),
                "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat(),
                "format": p.suffix.lstrip("."),
            })
    return files
