import asyncio
import re
import openpyxl
from datetime import datetime
from playwright.async_api import async_playwright

SEARCH_QUERY = [input("What do you want to scrape? ")]

MAX_RESULTS_PER_QUERY = 100
file = input("What should the output file name be? ")
OUTPUT_FILE = f"excel_sheets/{file}"

# -- here we put our scraping logic, we will be using asyncio because the program waits for external resources --

async def scrape_google_maps(query: str, max_results: int) -> list[dict]:
    results = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = await context.new_page()

        search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
        print(f"\nSearching: {query}")
        await page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(3000)

        results_panel = page.locator('div[role="feed"]')
        previous_count = 0

        # -- we will scroll first to load all listings before clicking anything --
        for scroll_attempt in range(15):
            listings = await page.locator('div[role="feed"] > div > div[jsaction]').all()
            current_count = len(listings)

            if current_count >= max_results:
                break
            if current_count == previous_count and scroll_attempt > 5:
                break

            previous_count = current_count
            await results_panel.evaluate("el => el.scrollBy(0, 2000)")
            await page.wait_for_timeout(2000)
            print(f"Scrolled - {current_count} listings found so far...")

        # -- then we would extract data from each listing --
        listings = await page.locator('div[role="feed"] > div > div[jsaction]').all()
        print(f"Processing {min(len(listings), max_results)} listings...")

        # -- deduplication --
        seen_in_scrape = set()  

        for listing in listings[:max_results]:
            try:
                await listing.click()
                await page.wait_for_timeout(2500)

                name = ""
                try:
                    name = await page.locator('h1.DUwDvf').inner_text(timeout=3000)
                except:
                    pass

                address = ""
                try:
                    address_el = page.locator('button[data-item-id="address"]')
                    if await address_el.count() > 0:
                        address = await address_el.inner_text(timeout=3000)
                except:
                    pass

                phone = ""
                try:
                    phone_el = page.locator('button[data-item-id^="phone"]')
                    if await phone_el.count() > 0:
                        phone = await phone_el.inner_text(timeout=3000)
                        phone = re.sub(r'[^\d\+\-\s]', '', phone).strip()
                except:
                    pass

                website = ""
                try:
                    web_el = page.locator('a[data-item-id="authority"]')
                    if await web_el.count() > 0:
                        website = await web_el.get_attribute("href", timeout=3000)
                except:
                    pass

                rating = ""
                try:
                    rating = await page.locator('div.F7nice span[aria-hidden="true"]').first.inner_text(timeout=2000)
                except:
                    pass

                category = ""
                try:
                    category = await page.locator('button.DkEaL').inner_text(timeout=2000)
                except:
                    pass

                if name:
                    dedup_key = (name.strip().lower(), phone.strip())
                    if dedup_key in seen_in_scrape:
                        print(f"Skipping duplicate: {name.strip()}")
                        continue
                    seen_in_scrape.add(dedup_key)

                    entry = {
                        "Name": name.strip(),
                        "Category": category.strip(),
                        "Address": address.strip(),
                        "Phone": phone.strip(),
                        "Website": website.strip(),
                        "Rating": rating.strip(),
                        "Search Query": query,
                    }
                    results.append(entry)
                    print(f"{name.strip()} | {phone.strip() or 'No phone'}")

            except Exception as e:
                continue

        await browser.close()
    return results

# -- a function to export the results to excel --
def save_to_excel(all_results: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    headers = ["Name", "Category", "Address", "Phone", "Website", "Rating", "Search Query"]

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, result in enumerate(all_results, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=result.get(header, ""))

    col_widths = [35, 25, 50, 18, 40, 10, 35]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    wb.save(filename)
    print(f"\nSaved {len(all_results)} records to '{filename}'")
    return len(all_results)

async def main():
    print("A tool to scrape listings from Google Map.")
    print(f"You started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    all_results = []

    for query in SEARCH_QUERY:
        results = await scrape_google_maps(query, MAX_RESULTS_PER_QUERY)
        all_results.extend(results)
        await asyncio.sleep(3)

    total = save_to_excel(all_results, OUTPUT_FILE)

    print(f"Done! {total} listings saved.")
    print(f"File: {OUTPUT_FILE}")

if __name__ == "__main__":
    asyncio.run(main())